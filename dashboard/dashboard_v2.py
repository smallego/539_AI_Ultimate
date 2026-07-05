# dashboard/dashboard_v2.py

from pathlib import Path
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.append(str(BASE_DIR / "dashboard"))

from dashboard_data import (
    read_latest_draw,
    read_weights_top10,
    read_latest_predictions,
    read_backtest,
    read_tuning,
)

OUTPUT_PATH = BASE_DIR / "dashboard" / "Dashboard_V2.xlsx"


def write_table(ws, start_row, title, headers, rows):
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = bold
    row = start_row + 1

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill

    row += 1

    for r in rows:
        for col, value in enumerate(r, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    return row + 1


def build_dashboard_v2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard V2"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=16)

    ws["A1"] = "539 AI Ultimate Dashboard V2"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    row = 3

    latest = read_latest_draw()
    if not latest.empty:
        r = latest.iloc[0]
        latest_rows = [[
            r["draw_no"],
            r["draw_date"],
            int(r["n1"]),
            int(r["n2"]),
            int(r["n3"]),
            int(r["n4"]),
            int(r["n5"]),
        ]]
        row = write_table(
            ws, row, "最新一期",
            ["期別", "日期", "N1", "N2", "N3", "N4", "N5"],
            latest_rows
        )

    predictions = read_latest_predictions()
    if not predictions.empty:
        pred_rows = []
        for _, r in predictions.iterrows():
            pred_rows.append([
                int(r["set_no"]),
                int(r["n1"]),
                int(r["n2"]),
                int(r["n3"]),
                int(r["n4"]),
                int(r["n5"]),
            ])

        row = write_table(
            ws, row, "最新預測 5 組",
            ["組別", "N1", "N2", "N3", "N4", "N5"],
            pred_rows
        )

    weights = read_weights_top10()
    if not weights.empty:
        weight_rows = [
            [int(r["number"]), round(float(r["final_weight"]), 4)]
            for _, r in weights.iterrows()
        ]
        row = write_table(
            ws, row, "信心號碼 TOP10",
            ["號碼", "權重"],
            weight_rows
        )

    tuning = read_tuning()
    if not tuning.empty:
        best = tuning.sort_values("roi", ascending=False).iloc[0]
        tuning_rows = [
            ["Weight", best["weight"]],
            ["Co-occurrence", best["co"]],
            ["Balance", best["balance"]],
            ["AI", best["ai"]],
            ["ROI", f"{best['roi']:.2f}%"],
            ["2碼率", f"{best['hit2']:.2f}%"],
            ["3碼率", f"{best['hit3']:.2f}%"],
            ["4碼率", f"{best['hit4']:.2f}%"],
        ]
        row = write_table(
            ws, row, "Auto Tuning 最佳權重",
            ["項目", "數值"],
            tuning_rows
        )

    backtest = read_backtest()
    if not backtest.empty:
        final = backtest.iloc[-1]
        hit2 = (backtest["best_match"] >= 2).mean() * 100
        hit3 = (backtest["best_match"] >= 3).mean() * 100
        hit4 = (backtest["best_match"] >= 4).mean() * 100

        backtest_rows = [
            ["回測期數", len(backtest)],
            ["累積 ROI", f"{final['cumulative_roi']:.2f}%"],
            ["2碼率", f"{hit2:.2f}%"],
            ["3碼率", f"{hit3:.2f}%"],
            ["4碼率", f"{hit4:.2f}%"],
        ]

        row = write_table(
            ws, row, "回測績效",
            ["項目", "數值"],
            backtest_rows
        )

        ws_data = wb.create_sheet("BacktestData")
        headers = ["draw_date", "best_match", "cumulative_roi"]

        for col, h in enumerate(headers, start=1):
            ws_data.cell(row=1, column=col, value=h)

        for idx, (_, r) in enumerate(backtest.iterrows(), start=2):
            ws_data.cell(row=idx, column=1, value=str(r["draw_date"]))
            ws_data.cell(row=idx, column=2, value=int(r["best_match"]))
            ws_data.cell(row=idx, column=3, value=float(r["cumulative_roi"]))

        chart = LineChart()
        chart.title = "Cumulative ROI Trend"
        chart.y_axis.title = "ROI %"
        chart.x_axis.title = "Draw"

        data = Reference(ws_data, min_col=3, min_row=1, max_row=len(backtest) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.height = 8
        chart.width = 18

        ws.add_chart(chart, "D4")

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 18

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center")

    wb.save(OUTPUT_PATH)

    print("===================================")
    print("Dashboard V2 已建立完成")
    print(f"輸出位置：{OUTPUT_PATH}")
    print("===================================")


if __name__ == "__main__":
    build_dashboard_v2()