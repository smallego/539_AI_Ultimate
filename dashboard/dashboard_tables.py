from openpyxl.styles import Font, PatternFill

bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9EAF7")


def write_table(ws, start_row, title, headers, rows):

    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = bold

    row = start_row + 1

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = bold
        c.fill = header_fill

    row += 1

    for r in rows:
        for col, value in enumerate(r, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    return row + 1