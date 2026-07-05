from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from dashboard.dashboard_tables import write_table
from dashboard.dashboard_stats import analytics_tables, period_comparison
from dashboard.dashboard_charts import add_bar_chart, add_line_chart, add_multi_line_chart, add_pie_chart
from dashboard.dashboard_trend import trend_data

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_PATH = BASE_DIR / "dashboard" / "Dashboard_V3.xlsx"


def style_sheet(ws, max_col=10):
    for c in range(1, max_col + 1):
        ws.column_dimensions[chr(64 + c)].width = 18

    for rows in ws.iter_rows():
        for cell in rows:
            cell.alignment = Alignment(horizontal="center")


def write_title(ws, title, merge="A1:F1"):
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=16)

    ws["A1"] = title
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells(merge)


def build_analytics_sheet(wb):
    ws = wb.active
    ws.title = "Analytics"
    write_title(ws, "539 AI Analytics Dashboard V4", "A1:F1")

    row = 3
    analytics, hot, cold = analytics_tables(last_n=100)

    row = write_table(ws, row, "Analytics (最近100期)", ["項目", "數值"], analytics)

    compare_rows = period_comparison()
    row = write_table(
        ws,
        row,
        "區間比較",
        ["區間", "期數", "平均和值", "平均跨度", "最熱號", "最冷號"],
        compare_rows,
    )

    hot_start = row
    row = write_table(ws, row, "Hot Number TOP10", ["號碼", "次數"], hot)
    add_bar_chart(ws, "Hot Number", hot_start)

    cold_start = row
    row = write_table(ws, row, "Cold Number TOP10", ["號碼", "次數"], cold)
    add_bar_chart(ws, "Cold Number", cold_start)

    # Pie chart source area
    pie_row = row + 1
    ws.cell(row=pie_row, column=1, value="奇偶")
    ws.cell(row=pie_row, column=2, value="數量")
    stats_map = {name: value for name, value in analytics}
    ws.cell(row=pie_row + 1, column=1, value="奇數")
    ws.cell(row=pie_row + 1, column=2, value=stats_map.get("奇數", 0))
    ws.cell(row=pie_row + 2, column=1, value="偶數")
    ws.cell(row=pie_row + 2, column=2, value=stats_map.get("偶數", 0))
    add_pie_chart(ws, "奇偶比例", 1, 2, pie_row + 1, 2, "E38")

    ws.cell(row=pie_row + 4, column=1, value="大小")
    ws.cell(row=pie_row + 4, column=2, value="數量")
    ws.cell(row=pie_row + 5, column=1, value="小號")
    ws.cell(row=pie_row + 5, column=2, value=stats_map.get("小號(1~20)", 0))
    ws.cell(row=pie_row + 6, column=1, value="大號")
    ws.cell(row=pie_row + 6, column=2, value=stats_map.get("大號(21~39)", 0))
    add_pie_chart(ws, "大小比例", 1, 2, pie_row + 5, 2, "E54")

    style_sheet(ws, 10)


def build_trend_sheet(wb):
    ws = wb.create_sheet("Trend")
    write_title(ws, "Trend Analysis 最近100期", "A1:H1")

    df = trend_data(last_n=100)
    headers = list(df.columns)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=header).font = Font(bold=True)

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    max_row = len(df) + 3
    add_line_chart(ws, "和值趨勢", data_col=3, anchor="J3", min_row=3, max_row=max_row, category_col=1)
    add_line_chart(ws, "跨度趨勢", data_col=4, anchor="J20", min_row=3, max_row=max_row, category_col=1)
    add_multi_line_chart(ws, "奇偶趨勢", data_cols=[5, 6], anchor="J37", min_row=3, max_row=max_row, category_col=1)
    add_multi_line_chart(ws, "大小趨勢", data_cols=[7, 8], anchor="J54", min_row=3, max_row=max_row, category_col=1)

    style_sheet(ws, 14)


def build_dashboard():
    wb = Workbook()
    build_analytics_sheet(wb)
    build_trend_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Dashboard 建立完成：{OUTPUT_PATH}")


if __name__ == "__main__":
    build_dashboard()
