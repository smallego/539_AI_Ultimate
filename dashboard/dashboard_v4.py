from pathlib import Path
import sys

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from dashboard.dashboard_tables import write_table
from dashboard.dashboard_stats import analytics_tables
from dashboard.dashboard_charts import add_bar_chart, add_line_chart
from dashboard.dashboard_trend import trend_data

OUTPUT_PATH = BASE_DIR / "dashboard" / "Dashboard_V4.xlsx"


def setup_title(ws, title):
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"] = title
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")


def center_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")


def autosize(ws, max_col=10, width=16):
    for c in range(1, max_col + 1):
        ws.column_dimensions[chr(64 + c)].width = width


def build_analytics_sheet(wb):
    ws = wb.active
    ws.title = "Analytics"
    setup_title(ws, "539 AI Analytics Dashboard V4")

    row = 3
    analytics, hot, cold = analytics_tables()

    row = write_table(
        ws,
        row,
        "Analytics (最近100期)",
        ["項目", "數值"],
        analytics,
    )

    hot_start = row
    row = write_table(
        ws,
        row,
        "Hot Number TOP10",
        ["號碼", "次數"],
        hot,
    )
    add_bar_chart(ws, "Hot Number TOP10", hot_start)

    cold_start = row
    row = write_table(
        ws,
        row,
        "Cold Number TOP10",
        ["號碼", "次數"],
        cold,
    )
    add_bar_chart(ws, "Cold Number TOP10", cold_start)

    autosize(ws, max_col=8, width=18)
    center_cells(ws)


def build_trend_sheet(wb):
    trend = trend_data()
    ws = wb.create_sheet("Trend")
    setup_title(ws, "Trend Analysis 最近100期")

    headers = list(trend.columns)
    start_row = 3

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for idx, rowdata in enumerate(trend.itertuples(index=False), start=start_row + 1):
        for col, value in enumerate(rowdata, start=1):
            ws.cell(row=idx, column=col, value=value)

    max_row = start_row + len(trend)

    add_line_chart(ws, "和值趨勢", 1, 1, start_row, max_row, "H3", "和值")
    add_line_chart(ws, "跨度趨勢", 2, 2, start_row, max_row, "H20", "跨度")
    add_line_chart(ws, "奇偶趨勢", 3, 4, start_row, max_row, "H37", "個數")
    add_line_chart(ws, "大小趨勢", 5, 6, start_row, max_row, "H54", "個數")

    autosize(ws, max_col=14, width=14)
    center_cells(ws)


def build_dashboard_v4():
    wb = Workbook()
    build_analytics_sheet(wb)
    build_trend_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Dashboard V4 建立完成：{OUTPUT_PATH}")


# compatibility alias for existing GUI/main naming
build_dashboard = build_dashboard_v4


if __name__ == "__main__":
    build_dashboard_v4()
