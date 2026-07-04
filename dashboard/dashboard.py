# dashboard/dashboard.py

import pandas as pd
import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, Reference
BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "database" / "history.db"
REPORT_PATH = BASE_DIR / "reports" / "backtest_result.csv"
OUTPUT_PATH = BASE_DIR / "dashboard" / "Dashboard.xlsx"


def read_backtest():
    if not REPORT_PATH.exists():
        return None
    return pd.read_csv(REPORT_PATH)


def read_weights():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""
        SELECT number, final_weight
        FROM model_weights
        ORDER BY final_weight DESC
        LIMIT 10
    """, conn)
    conn.close()
    return df


def read_latest_predictions():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""
        SELECT set_no, n1, n2, n3, n4, n5
        FROM prediction_history
        ORDER BY id DESC
        LIMIT 5
    """, conn)
    conn.close()
    return df.sort_values("set_no")


def build_dashboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    ws["A1"] = "539 AI Ultimate Dashboard"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")

    backtest = read_backtest()
    weights = read_weights()
    predictions = read_latest_predictions()

    row = 3

    ws[f"A{row}"] = "模型績效"
    ws[f"A{row}"].font = bold
    row += 1

    if backtest is not None and len(backtest) > 0:
        final = backtest.iloc[-1]
        test_count = len(backtest)
        roi = final["cumulative_roi"]
        hit2 = (backtest["best_match"] >= 2).mean() * 100
        hit3 = (backtest["best_match"] >= 3).mean() * 100
        hit4 = (backtest["best_match"] >= 4).mean() * 100

        metrics = [
            ("回測期數", test_count),
            ("累積 ROI", f"{roi:.2f}%"),
            ("2碼以上命中率", f"{hit2:.2f}%"),
            ("3碼以上命中率", f"{hit3:.2f}%"),
            ("4碼以上命中率", f"{hit4:.2f}%"),
        ]

        for name, value in metrics:
            ws[f"A{row}"] = name
            ws[f"B{row}"] = value
            row += 1
    else:
        ws[f"A{row}"] = "尚未產生 backtest_result.csv"
        row += 1

    row += 2
    ws[f"A{row}"] = "最新預測 5 組"
    ws[f"A{row}"].font = bold
    row += 1

    headers = ["組別", "N1", "N2", "N3", "N4", "N5"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = bold

    row += 1

    for _, r in predictions.iterrows():
        ws.cell(row=row, column=1, value=int(r["set_no"]))
        for i, col in enumerate(["n1", "n2", "n3", "n4", "n5"], start=2):
            ws.cell(row=row, column=i, value=int(r[col]))
        row += 1

    row += 2
    ws[f"A{row}"] = "信心號碼 TOP10"
    ws[f"A{row}"].font = bold
    row += 1

    ws[f"A{row}"] = "號碼"
    ws[f"B{row}"] = "權重"
    ws[f"A{row}"].fill = header_fill
    ws[f"B{row}"].fill = header_fill
    ws[f"A{row}"].font = bold
    ws[f"B{row}"].font = bold
    row += 1

    for _, r in weights.iterrows():
        ws[f"A{row}"] = int(r["number"])
        ws[f"B{row}"] = round(float(r["final_weight"]), 4)
        row += 1

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center")

    # ===== 建立 BacktestData 工作表 =====
    if backtest is not None and len(backtest) > 0:
        ws_data = wb.create_sheet("BacktestData")

        data_headers = [
            "draw_date",
            "best_match",
            "period_roi",
            "cumulative_roi",
        ]

        for col, h in enumerate(data_headers, start=1):
            ws_data.cell(row=1, column=col, value=h)

        for idx, (_, r) in enumerate(backtest.iterrows(), start=2):
            ws_data.cell(row=idx, column=1, value=str(r["draw_date"]))
            ws_data.cell(row=idx, column=2, value=int(r["best_match"]))
            ws_data.cell(row=idx, column=3, value=float(r["period_roi"]))
            ws_data.cell(row=idx, column=4, value=float(r["cumulative_roi"]))

        # ROI 趨勢圖
        chart_roi = LineChart()
        chart_roi.title = "Cumulative ROI Trend"
        chart_roi.y_axis.title = "ROI %"
        chart_roi.x_axis.title = "Draw"

        data = Reference(ws_data, min_col=4, min_row=1, max_row=len(backtest) + 1)
        chart_roi.add_data(data, titles_from_data=True)
        chart_roi.height = 8
        chart_roi.width = 18

        ws.add_chart(chart_roi, "D4")

        # Best Match 走勢圖
        chart_match = LineChart()
        chart_match.title = "Best Match Trend"
        chart_match.y_axis.title = "Match Count"
        chart_match.x_axis.title = "Draw"

        data_match = Reference(ws_data, min_col=2, min_row=1, max_row=len(backtest) + 1)
        chart_match.add_data(data_match, titles_from_data=True)
        chart_match.height = 8
        chart_match.width = 18

        ws.add_chart(chart_match, "D20")
    wb.save(OUTPUT_PATH)

    print("===================================")
    print("Dashboard 已建立完成")
    print(f"輸出位置：{OUTPUT_PATH}")
    print("===================================")


if __name__ == "__main__":
    build_dashboard()