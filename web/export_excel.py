from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from web.services import (
    ai_signal,
    dashboard_pro,
    decision_center,
    latest_predictions,
    prediction_history,
    prediction_performance,
    prediction_result_analysis,
    strategy_lab,
    strategy_optimizer,
    system_info,
)


HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="17365D", bold=True, size=16)


def export_time():
    return datetime.now()


def excel_filename():
    return f"539_AI_Ultimate_Report_{export_time().strftime('%Y%m%d_%H%M%S')}.xlsx"


def number_text(numbers):
    if not numbers:
        return "-"
    return " ".join(str(number).zfill(2) for number in numbers)


def write_title(sheet, version):
    now = export_time()
    rows = [
        ["539 AI Ultimate"],
        ["Professional Report"],
        ["Export Time", now],
        ["Version", version or "-"],
        [],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = TITLE_FONT
    sheet["A2"].font = Font(bold=True)
    sheet["B3"].number_format = "yyyy-mm-dd hh:mm:ss"


def style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def append_table(sheet, title, headers, rows):
    sheet.append([title])
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=13)
    sheet.append(headers)
    header_row = sheet[sheet.max_row]
    style_header(header_row)
    for row in rows:
        sheet.append(row)
    sheet.append([])


def apply_sheet_style(sheet):
    sheet.freeze_panes = "A7"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        width = 12
        for cell in sheet[letter]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 42))
        sheet.column_dimensions[letter].width = width

    for row in sheet.iter_rows():
        labels = [str(cell.value).lower() for cell in row if cell.value is not None]
        if any("roi" == label or "roi" in label for label in labels):
            continue


def set_number_formats(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            header = str(sheet.cell(row=6, column=cell.column).value or "").lower()
            nearby = " ".join(str(sheet.cell(row=max(1, cell.row - offset), column=cell.column).value or "").lower() for offset in range(1, 4))
            label = f"{header} {nearby}"
            if "roi" in label and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
                if abs(cell.value) > 1:
                    cell.value = cell.value / 100
            elif "score" in label and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
            elif "prize" in label and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def latest_draw_summary(dashboard):
    summary = dashboard.get("todaySummary") or {}
    return summary


def dashboard_sheet(workbook, version):
    sheet = workbook.create_sheet("Dashboard")
    write_title(sheet, version)
    dashboard = dashboard_pro()
    signal = ai_signal()
    system = system_info()
    summary = latest_draw_summary(dashboard)
    decision = dashboard.get("decisionOverview") or {}
    strategy = dashboard.get("strategyOverview") or {}

    rows = [
        ["Latest Draw", summary.get("latestDraw")],
        ["Latest Date", summary.get("drawDate")],
        ["Decision Score", decision.get("decisionScore")],
        ["AI Score", (decision.get("components") or {}).get("aiScore")],
        ["Buy Signal", signal.get("signal")],
        ["Risk", signal.get("risk")],
        ["Confidence", signal.get("confidence")],
        ["Best Strategy", strategy.get("bestStrategy")],
        ["Prediction Count", system.get("predictionCount")],
        ["History Count", system.get("databaseCount")],
        ["Learning Count", system.get("learningCount")],
    ]
    append_table(sheet, "Dashboard", ["Metric", "Value"], rows)
    apply_sheet_style(sheet)
    return sheet


def prediction_sheet(workbook, version):
    sheet = workbook.create_sheet("Prediction")
    write_title(sheet, version)
    analysis = prediction_result_analysis()
    result_map = {row.get("id"): row for row in analysis.get("rows") or []}
    rows = []
    for prediction in latest_predictions():
        result = result_map.get(prediction.get("id"), {})
        rows.append([
            prediction.get("set_no"),
            number_text(prediction.get("numbers")),
            prediction.get("ai_score"),
            prediction.get("final_score"),
            result.get("hits"),
            number_text(result.get("matched_numbers")),
            result.get("prize"),
            result.get("status"),
            prediction.get("created_at"),
        ])
    append_table(
        sheet,
        "Prediction",
        ["Set", "Numbers", "AI Score", "Final Score", "Hit Count", "Matched Numbers", "Prize", "Status", "Prediction Time"],
        rows,
    )
    apply_sheet_style(sheet)
    set_number_formats(sheet)
    return sheet


def decision_sheet(workbook, version):
    sheet = workbook.create_sheet("Decision")
    write_title(sheet, version)
    decision = decision_center()
    reasons = decision.get("reasons") or {}
    trend = " / ".join(map(str, reasons.get("Trend") or []))
    rows = [
        ["Decision Score", decision.get("decisionScore")],
        ["Recommendation", decision.get("recommendation")],
        ["Confidence", decision.get("confidence")],
        ["Risk", decision.get("risk")],
        ["Hot Numbers", number_text(reasons.get("Hot Number"))],
        ["Cold Numbers", number_text(reasons.get("Cold Number"))],
        ["Trend", trend],
        ["Reasons", " | ".join(f"{key}: {' / '.join(map(str, value or []))}" for key, value in reasons.items())],
    ]
    append_table(sheet, "Decision", ["Metric", "Value"], rows)
    apply_sheet_style(sheet)
    return sheet


def strategy_sheet(workbook, version):
    sheet = workbook.create_sheet("Strategy")
    write_title(sheet, version)
    lab = strategy_lab()
    optimizer = strategy_optimizer()
    optimizer_map = {row.get("name"): row for row in optimizer.get("strategies") or []}
    rows = []
    for index, strategy in enumerate(lab.get("ranking") or [], start=1):
        optimized = optimizer_map.get(strategy.get("name"), {})
        rows.append([
            strategy.get("name"),
            strategy.get("roi"),
            strategy.get("hit2Rate"),
            strategy.get("risk"),
            optimized.get("confidence"),
            optimized.get("overallScore") or strategy.get("score"),
            index,
        ])
    append_table(
        sheet,
        "Strategy",
        ["Strategy Name", "ROI", "Hit Rate", "Risk", "Confidence", "Overall Score", "Rank"],
        rows,
    )
    apply_sheet_style(sheet)
    set_number_formats(sheet)
    return sheet


def performance_sheet(workbook, version):
    sheet = workbook.create_sheet("Performance")
    write_title(sheet, version)
    rows = []
    for row in prediction_performance().get("recentRows") or []:
        prize = row.get("prize") or 0
        roi = ((prize - 50) / 50) if prize is not None else None
        rows.append([
            row.get("created_at"),
            row.get("hits"),
            prize,
            roi,
            row.get("ai_score"),
            row.get("final_score"),
        ])
    append_table(
        sheet,
        "Prediction History",
        ["Date", "Hits", "Prize", "ROI", "AI Score", "Final Score"],
        rows,
    )
    apply_sheet_style(sheet)
    set_number_formats(sheet)
    return sheet


def build_excel_report():
    version = system_info().get("version")
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    dashboard_sheet(workbook, version)
    prediction_sheet(workbook, version)
    decision_sheet(workbook, version)
    strategy_sheet(workbook, version)
    performance_sheet(workbook, version)

    for sheet in workbook.worksheets:
        sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
