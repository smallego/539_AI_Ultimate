from datetime import datetime
from io import BytesIO

from web.services import (
    ai_signal,
    backtest_center,
    dashboard_pro,
    decision_center,
    latest_predictions,
    prediction_history,
    prediction_result_analysis,
    strategy_lab,
    strategy_optimizer,
)


EXPORT_TITLES = {
    "dashboard": "539 AI Dashboard",
    "prediction": "539 AI Prediction",
    "decision": "539 AI Decision",
    "strategy": "539 AI Strategy",
}


def timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def export_filename(report_type, extension):
    name = {
        "dashboard": "Dashboard",
        "prediction": "Prediction",
        "decision": "Decision",
        "strategy": "Strategy",
    }.get(report_type, "Dashboard")
    return f"539_AI_{name}_{timestamp()}.{extension}"


def number_text(numbers):
    if not numbers:
        return "-"
    return " ".join(str(number).zfill(2) for number in numbers)


def flatten_rows(rows):
    clean = []
    for row in rows:
        clean.append([
            ", ".join(map(str, value)) if isinstance(value, list) else value
            for value in row
        ])
    return clean


def dashboard_sections():
    data = dashboard_pro()
    signal = ai_signal()
    summary = data.get("todaySummary") or {}
    decision = data.get("decisionOverview") or {}
    strategy = data.get("strategyOverview") or {}
    performance = data.get("performance") or {}

    return [
        ("AI Buy Signal", [
            ["Signal", signal.get("signal")],
            ["Score", signal.get("score")],
            ["Confidence", signal.get("confidence")],
            ["Risk", signal.get("risk")],
            ["Reasons", " / ".join(signal.get("reason") or [])],
        ]),
        ("Dashboard Summary", [
            ["Latest Draw", summary.get("latestDraw")],
            ["Latest Draw Date", summary.get("drawDate")],
            ["Latest Numbers", number_text(summary.get("latestNumbers"))],
            ["Best Pick", f"Set {summary.get('bestPick', {}).get('set_no', '-')}"],
            ["Decision Score", decision.get("decisionScore")],
            ["Recommendation", decision.get("recommendation")],
            ["Best Strategy", strategy.get("bestStrategy")],
            ["Draw Status", summary.get("drawStatus")],
        ]),
        ("Strategy Ranking", [
            ["Rank", "Strategy", "Score", "Risk"],
            *[
                [index + 1, row.get("name"), row.get("score"), row.get("risk")]
                for index, row in enumerate(strategy.get("ranking") or [])
            ],
        ]),
        ("Performance", [
            ["Average API Time", performance.get("averageApiTime")],
            ["Cache Hit Rate", performance.get("cacheHitRate")],
            ["Database Count", performance.get("databaseCount")],
            ["Prediction Count", performance.get("predictionCount")],
            ["Learning Count", performance.get("learningCount")],
        ]),
    ]


def prediction_sections():
    dashboard = dashboard_pro()
    analysis = prediction_result_analysis()
    latest = latest_predictions()
    history = prediction_history(limit=50)
    result_map = {row.get("id"): row for row in analysis.get("rows") or []}
    latest_draw = (dashboard.get("todaySummary") or {})

    return [
        ("Latest Draw", [
            ["Draw No", latest_draw.get("latestDraw")],
            ["Draw Date", latest_draw.get("drawDate")],
            ["Numbers", number_text(latest_draw.get("latestNumbers"))],
        ]),
        ("Current Prediction", [
            ["Set", "Numbers", "Final Score", "AI Score"],
            *[
                [row.get("set_no"), number_text(row.get("numbers")), row.get("final_score"), row.get("ai_score")]
                for row in latest
            ],
        ]),
        ("Result Analysis", [
            ["Set", "Numbers", "Draw Numbers", "Hits", "Matched", "Hit Rate", "Prize Level", "Prize", "Status"],
            *[
                [
                    row.get("set_no"),
                    number_text(row.get("numbers")),
                    number_text(row.get("draw_numbers")),
                    row.get("hits"),
                    number_text(row.get("matched_numbers")),
                    row.get("hit_rate"),
                    row.get("prize_level"),
                    row.get("prize"),
                    row.get("status"),
                ]
                for row in analysis.get("rows") or []
            ],
        ]),
        ("Prediction History", [
            ["Time", "Set", "Numbers", "Hits", "Prize", "Final", "AI"],
            *[
                [
                    row.get("created_at"),
                    row.get("set_no"),
                    number_text(row.get("numbers")),
                    result_map.get(row.get("id"), {}).get("hits"),
                    result_map.get(row.get("id"), {}).get("prize"),
                    row.get("final_score"),
                    row.get("ai_score"),
                ]
                for row in history
            ],
        ]),
    ]


def decision_sections():
    signal = ai_signal()
    decision = decision_center()
    predictions = latest_predictions()
    reasons = decision.get("reasons") or {}

    return [
        ("AI Signal", [
            ["Signal", signal.get("signal")],
            ["Score", signal.get("score")],
            ["Confidence", signal.get("confidence")],
            ["Risk", signal.get("risk")],
        ]),
        ("Decision", [
            ["Decision Score", decision.get("decisionScore")],
            ["Recommendation", decision.get("recommendation")],
            ["Confidence", decision.get("confidence")],
            ["Risk", decision.get("risk")],
        ]),
        ("Current Prediction", [
            ["Set", "Numbers", "Final Score", "AI Score"],
            *[
                [row.get("set_no"), number_text(row.get("numbers")), row.get("final_score"), row.get("ai_score")]
                for row in predictions
            ],
        ]),
        ("Reasons", [
            ["Type", "Reason"],
            *[
                [key, " / ".join(map(str, values or []))]
                for key, values in reasons.items()
            ],
        ]),
    ]


def strategy_sections():
    lab = strategy_lab()
    optimizer = strategy_optimizer()
    recommendation = optimizer.get("recommendation") or {}

    return [
        ("Strategy Ranking", [
            ["Rank", "Strategy", "ROI", "Hit2", "Hit3", "Risk", "Score"],
            *[
                [
                    index + 1,
                    row.get("name"),
                    row.get("roi"),
                    row.get("hit2Rate"),
                    row.get("hit3Rate"),
                    row.get("risk"),
                    row.get("score"),
                ]
                for index, row in enumerate(lab.get("ranking") or [])
            ],
        ]),
        ("Optimizer", [
            ["Best Strategy", recommendation.get("strategy")],
            ["Overall Score", recommendation.get("overallScore")],
            ["Risk", recommendation.get("risk")],
            ["Confidence", recommendation.get("confidence")],
        ]),
        ("Reasons", [
            ["Reason"],
            *[[reason] for reason in recommendation.get("reasons") or []],
        ]),
    ]


def report_sections(report_type):
    builders = {
        "dashboard": dashboard_sections,
        "prediction": prediction_sections,
        "decision": decision_sections,
        "strategy": strategy_sections,
    }
    return builders.get(report_type, dashboard_sections)()


def build_xlsx(report_type):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXPORT_TITLES.get(report_type, "539 AI")[:31]
    row_index = 1

    sheet.cell(row=row_index, column=1, value=EXPORT_TITLES.get(report_type, "539 AI Report"))
    sheet.cell(row=row_index, column=1).font = Font(bold=True, size=16)
    row_index += 2

    for title, rows in report_sections(report_type):
        sheet.cell(row=row_index, column=1, value=title)
        sheet.cell(row=row_index, column=1).font = Font(bold=True, size=13)
        sheet.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor="DCE5EF")
        row_index += 1
        for row in flatten_rows(rows):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
            row_index += 1
        row_index += 1

    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_pdf(report_type):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(EXPORT_TITLES.get(report_type, "539 AI Report"), styles["Title"]), Spacer(1, 12)]

    for title, rows in report_sections(report_type):
        story.append(Paragraph(title, styles["Heading2"]))
        table_rows = flatten_rows(rows[:60])
        table = Table(table_rows, repeatRows=1 if len(table_rows) > 1 else 0)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE5EF")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AAB7C4")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.extend([table, Spacer(1, 12)])

    document.build(story)
    output.seek(0)
    return output
